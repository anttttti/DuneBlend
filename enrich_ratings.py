#!/usr/bin/env python3
"""
Fetch card ratings from dunecardshub.com and write them to the Excel spreadsheet.
Adds columns: DCH Rating (1-5), DCH Votes, DCH Tier (S/A/B/C/D)
"""
import requests
import json
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).parent / "Dune_Imperium_Card_Inventory.xlsx"
BASE_URL = "https://dunecardshub.com"

EXPANSION_MAP = {
    "Dune: Imperium": ["Base", "Imperium"],
    "Rise of Ix": ["Rise of Ix", "Ix"],
    "Immortality": ["Immortality"],
    "Uprising": ["Uprising"],
    "Bloodlines": ["Bloodlines"],
    "Promo": ["Promo"],
}

# Map API type → Excel sheet name
TYPE_TO_SHEET = {
    "imperium": "Imperium",
    "intrigue": "Intrigue",
}

# Map rating number → tier letter (5=S, 4=A, 3=B, 2=C, 1=D)
RATING_TO_TIER = {5: "S", 4: "A", 3: "B", 2: "C", 1: "D"}

# Name corrections: maps our Excel names (lowercased) → API names (lowercased)
NAME_CORRECTIONS = {
    "ruthless leadership": "ruthless leadeship",       # API typo
    "full-scale assault": "full-scale assult",         # API typo
    "to the victor…": "to the victory...",
    "change allegiences": "change allegiances",        # Excel typo
    "ornithopter": "ornitopter",                       # API alternate spelling
    "corner the market": "corner the marker",          # API alternate spelling
}


def fetch_all_ratings():
    headers = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36"}
    all_cards = []
    page = 1
    while True:
        r = requests.get(
            f"{BASE_URL}/api/cards/rating?limit=100&page={page}",
            headers=headers, verify=False, timeout=20
        )
        data = r.json()
        all_cards.extend(data["data"])
        print(f"  Fetched page {page}/{data['pagination']['totalPages']}")
        if not data["pagination"]["hasNextPage"]:
            break
        page += 1
    print(f"  Total cards fetched: {len(all_cards)}")
    return all_cards


def build_lookup(all_cards):
    """Build lookup: (normalized_name, api_type) -> rating data"""
    lookup = {}
    for c in all_cards:
        card = c["card"]
        name = card["name"].strip().lower()
        api_type = card["type"]
        expansion = card["expansion"] or ""
        key = (name, api_type, expansion)
        lookup[key] = {
            "rating": c["rating"],
            "votes": c["votesCount"],
            "tier": RATING_TO_TIER.get(round(c["rating"]) if c["rating"] else 0, ""),
        }
        # Also store without expansion for fallback matching
        key2 = (name, api_type, "")
        if key2 not in lookup:
            lookup[key2] = lookup[key]
    return lookup


def add_columns_to_sheet(ws, sheet_name, lookup, api_type):
    """Add DCH Rating, DCH Votes, DCH Tier columns to a sheet."""
    headers = [cell.value for cell in ws[1]]

    # Determine column indices for new columns (add if not present)
    new_cols = ["DCH Rating", "DCH Votes", "DCH Tier"]
    col_indices = {}
    for col_name in new_cols:
        if col_name in headers:
            col_indices[col_name] = headers.index(col_name) + 1
        else:
            next_col = len(headers) + 1
            ws.cell(row=1, column=next_col, value=col_name)
            col_indices[col_name] = next_col
            headers.append(col_name)
            print(f"  Added column '{col_name}' at position {next_col}")

    name_col_idx = 0  # First column is the name

    # Find Source column index
    source_col_idx = headers.index("Source") if "Source" in headers else None

    matched = 0
    unmatched = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        raw_name = row[name_col_idx].value
        if not raw_name:
            continue
        card_name = str(raw_name).strip().lower()
        source_val = str(row[source_col_idx].value).strip() if source_col_idx is not None and row[source_col_idx].value else ""

        # Map source to API expansion
        api_expansions = []
        for api_exp, excel_sources in EXPANSION_MAP.items():
            if source_val in excel_sources:
                api_expansions.append(api_exp)

        # Apply name corrections if needed
        corrected_name = NAME_CORRECTIONS.get(card_name, card_name)

        # Try to match with expansion, then without
        rating_data = None
        for name_try in [corrected_name, card_name]:
            for api_exp in api_expansions:
                rating_data = lookup.get((name_try, api_type, api_exp))
                if rating_data:
                    break
            if not rating_data:
                rating_data = lookup.get((name_try, api_type, ""))
            if rating_data:
                break

        if rating_data:
            ws.cell(row=row_idx, column=col_indices["DCH Rating"]).value = rating_data["rating"]
            ws.cell(row=row_idx, column=col_indices["DCH Votes"]).value = rating_data["votes"]
            ws.cell(row=row_idx, column=col_indices["DCH Tier"]).value = rating_data["tier"]
            matched += 1
        else:
            unmatched.append(f"{raw_name!r} [{source_val}]")

    print(f"  {sheet_name}: matched {matched}, unmatched {len(unmatched)}")
    if unmatched:
        print(f"  Unmatched cards:")
        for u in unmatched:
            print(f"    {u}")


def main():
    print("Fetching ratings from dunecardshub.com...")
    all_cards = fetch_all_ratings()
    lookup = build_lookup(all_cards)

    print(f"\nLoading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    for api_type, sheet_name in TYPE_TO_SHEET.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' not found, skipping")
            continue
        print(f"\nProcessing sheet: {sheet_name}")
        ws = wb[sheet_name]
        add_columns_to_sheet(ws, sheet_name, lookup, api_type)

    wb.save(EXCEL_PATH)
    print(f"\nSaved: {EXCEL_PATH}")


if __name__ == "__main__":
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    main()
