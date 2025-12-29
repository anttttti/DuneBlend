"""
Dune Imperium Blend Builder - Multi-Resource Version
Handles all resource types: Imperium, Tleilax, Reserve, Intrigue, Tech, Contracts, etc.
"""
from flask import Flask, render_template, jsonify, request
from pathlib import Path
import openpyxl
import webbrowser
import threading

app = Flask(__name__)

# Global resource data
ALL_RESOURCES = {}


def load_all_resources_from_excel():
    """Load all resource types from all Excel worksheets."""
    excel_path = Path(__file__).parent / "Dune_Imperium_Card_Inventory.xlsx"

    if not excel_path.exists():
        raise FileNotFoundError(f"Could not find: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    all_resources = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        resources = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            row_dict = dict(zip(headers, row))

            # Get the name from first column
            name_col = headers[0]
            resource_name = str(row_dict.get(name_col, '')).strip()
            if not resource_name:
                continue
            # Normalize card names: replace (Base) with (Imperium)
            resource_name = resource_name.replace("(Base)", "(Imperium)")

            # Create resource object
            resource = {
                'resource_type': sheet_name.lower(),
                'name': resource_name
            }

            # Add all columns as properties
            for key, value in row_dict.items():
                if key and value is not None and key != name_col:
                    col_key = key.lower().replace(' ', '_').replace('-', '_')
                    resource[col_key] = str(value) if not isinstance(value, (int, float)) else value

            # Add source/set mapping for color coding
            source = resource.get('source', 'Imperium')
            # Normalize: replace 'Base' with 'Imperium'
            if source == 'Base':
                source = 'Imperium'
                resource['source'] = 'Imperium'
            card_set_mapping = {
                "Imperium": "base",
                "Base": "base",
                "Rise of Ix": "ix",
                "Ix": "ix",
                "Immortality": "immortality",
                "Uprising": "uprising",
                "Bloodlines": "bloodlines",
                "Promo": "promo"
            }
            resource['card_set'] = card_set_mapping.get(source, str(source).lower() if source else 'base')

            resources.append(resource)

        all_resources[sheet_name.lower()] = resources
        print(f"Loaded {len(resources)} items from {sheet_name}")

    return all_resources


@app.route('/')
def index():
    """Serve the main page."""
    return render_template('index.html')


@app.route('/api/resources')
def get_resources():
    """Return all resources grouped by type."""
    return jsonify(ALL_RESOURCES)


@app.route('/api/resources/<resource_type>')
def get_resources_by_type(resource_type):
    """Return resources of a specific type."""
    return jsonify(ALL_RESOURCES.get(resource_type.lower(), []))


@app.route('/api/blends')
def list_blends():
    """List all available blend files with their actual filenames."""
    blends_dir = Path(__file__).parent / 'blends'
    blends_dir.mkdir(exist_ok=True)

    blend_files = []
    for filepath in sorted(blends_dir.glob('*.md')):
        blend_files.append({
            'filename': filepath.name  # Keep actual filename as-is
        })

    return jsonify(blend_files)


@app.route('/api/blend/load/<filename>')
def load_blend(filename):
    """Load a blend file and parse all resource types from simplified format."""
    blends_dir = Path(__file__).parent / 'blends'
    filepath = blends_dir / filename

    if not filepath.exists():
        return jsonify({"success": False, "error": "Blend file not found"}), 404

    # Parse the markdown file
    resources_by_type = {}
    current_section = None

    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()

            # Check for main sections (## Resource Type)
            if line.startswith('## '):
                current_section = line[3:].strip()
                if current_section not in resources_by_type:
                    if current_section == 'Board':
                        resources_by_type[current_section] = {
                            'mainBoard': 'imperium',
                            'additionalBoards': []
                        }
                    else:
                        resources_by_type[current_section] = []

            # Parse Board metadata
            elif current_section == 'Board' and line.startswith('- '):
                clean_line = line.lstrip('- ').strip()
                if clean_line.startswith('Main Board:'):
                    main_board = clean_line.split(':', 1)[1].strip()
                    resources_by_type['Board']['mainBoard'] = main_board
                elif clean_line.startswith('Additional Boards:'):
                    boards_str = clean_line.split(':', 1)[1].strip()
                    additional = [b.strip() for b in boards_str.split(',')]
                    resources_by_type['Board']['additionalBoards'] = additional

            # Check for resource lines in format "- count× name" or legacy "count name"
            elif current_section and current_section != 'Board' and line and not line.startswith('**') and not line.startswith('*Generated'):
                # Remove leading "- " if present
                clean_line = line.lstrip('- ').strip()

                if not clean_line or clean_line.startswith('#'):
                    continue

                # Parse "count× name" or "count name" or just "name" format
                # Try with × first
                if '×' in clean_line:
                    parts = clean_line.split('×', 1)
                    if len(parts) == 2 and parts[0].strip().isdigit():
                        count = int(parts[0].strip())
                        name = parts[1].strip()
                        # Add this resource 'count' times
                        for _ in range(count):
                            resources_by_type[current_section].append(name)
                    else:
                        # Just a name with × in it
                        resources_by_type[current_section].append(clean_line)
                elif clean_line and clean_line[0].isdigit() and ' ' in clean_line:
                    # Fallback to space separator for backward compatibility "2 Card Name"
                    parts = clean_line.split(' ', 1)
                    if parts[0].isdigit():
                        count = int(parts[0])
                        name = parts[1].strip()
                        for _ in range(count):
                            resources_by_type[current_section].append(name)
                else:
                    # Just a name without count
                    resources_by_type[current_section].append(clean_line)

    return jsonify({"success": True, "resources": resources_by_type})


@app.route('/api/blend/save', methods=['POST'])
def save_blend():
    """Save blend with all resource types in simplified format."""
    data = request.json
    blend_name = data.get('name', 'Untitled Blend')
    resources_by_type = data.get('resources', {})

    blends_dir = Path(__file__).parent / 'blends'
    blends_dir.mkdir(exist_ok=True)

    # Build markdown
    md = f"# {blend_name}\n\n"

    # Handle Board section first if present
    if 'Board' in resources_by_type:
        board_data = resources_by_type['Board']
        md += f"## Board\n\n"
        md += f"- Main Board: {board_data.get('mainBoard', 'imperium')}\n"
        additional = board_data.get('additionalBoards', [])
        if additional:
            md += f"- Additional Boards: {', '.join(additional)}\n"
        md += "\n"

    # Count total items (excluding Board metadata)
    total_count = sum(len(items) for key, items in resources_by_type.items()
                     if key != 'Board' and isinstance(items, list))
    if total_count > 0:
        md += f"**Total Items:** {total_count}\n\n"

    # Add each resource type section
    for resource_type, items in resources_by_type.items():
        if resource_type == 'Board':
            continue  # Already handled
        if not items:
            continue

        md += f"## {resource_type}\n\n"

        # Count occurrences of each item with set/source
        item_counts = {}
        for item in items:
            # Use displayName for conflicts (includes #X suffix), objective for contracts, otherwise name
            if 'displayName' in item:
                item_name = item.get('displayName')
            else:
                item_name = item.get('objective') or item.get('name', 'Unknown')

            # Get the set/source
            item_source = item.get('source', 'Unknown')

            # Create unique key with name and source
            item_key = f"{item_name} ({item_source})"

            if item_key not in item_counts:
                item_counts[item_key] = 0
            item_counts[item_key] += 1

        # Sort by name and output in "- count× name (source)" format
        for item_key in sorted(item_counts.keys()):
            count = item_counts[item_key]
            if count == 1:
                md += f"- {item_key}\n"
            else:
                md += f"- {count}× {item_key}\n"

        md += "\n"

    md += "---\n*Generated by Dune Imperium Blend Builder*\n"

    filename = f"{blend_name}.md".replace(' ', '_').replace('/', '_')
    if not filename.endswith('.md'):
        filename += '.md'
    filepath = blends_dir / filename

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(md)

    return jsonify({"success": True, "filepath": str(filepath), "filename": filename})


def open_browser():
    """Open browser after short delay."""
    webbrowser.open('http://localhost:5000')


if __name__ == '__main__':
    print("Loading all resources from Excel...")
    ALL_RESOURCES = load_all_resources_from_excel()
    print(f"\nTotal resource types loaded: {len(ALL_RESOURCES)}")
    threading.Timer(1.5, open_browser).start()
    app.run(debug=False, port=5000)

