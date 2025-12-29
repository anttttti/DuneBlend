#!/usr/bin/env python3
"""
Create Merakon's House Blend with specific requirements:
- All Uprising Imperium cards
- All Uprising Conflicts
- All Uprising Contracts
- All Uprising Starter cards
- All Uprising Leaders + specific leaders from other sets
"""
import openpyxl
from pathlib import Path
from collections import Counter

def create_merakon_blend():
    excel_path = Path(__file__).parent / "Dune_Imperium_Card_Inventory.xlsx"
    blends_dir = Path(__file__).parent / "blends"
    blends_dir.mkdir(exist_ok=True)

    print(f"Loading {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Resource type configurations
    resource_sheets = {
        'Imperium': 'Imperium Cards',
        'Intrigue': 'Intrigue Cards',
        'Tleilax': 'Tleilax Cards',
        'Reserve': 'Reserve Cards',
        'Tech': 'Tech Tiles',
        'Contracts': 'Contracts',
        'Leader': 'Leaders',
        'Sardaukar': 'Sardaukar',
        'Starter': 'Starter Cards',
        'Conflict': 'Conflict Cards'
    }

    merakon_resources = {sheet: [] for sheet in resource_sheets.values()}

    # Specific leaders to add manually (not in Excel or need to be ensured)
    manual_leaders = {
        '"Princess" Yuna Moritani': 'Rise of Ix',
        'Glossu "The Beast" Rabban': 'Imperium'
    }

    # All Uprising leaders to include
    uprising_leaders = [
        "Feyd-Rautha Harkonnen",
        "Gurney Halleck",
        "Lady Amber Metulli",
        "Lady Jessica",
        "Lady Margot Fenring",
        "Muad'Dib",
        "Princess Irulan",
        "Shaddam Corrino IV",
        "Staban Tuek"
    ]

    # Specific leaders from other sets
    other_leaders = [
        "Archduke Armand Ecaz",
        "Baron Vladimir Harkonnen",
        "Count Ilban Richese",
        "Countess Ariana Thorvald",
        "Earl Memnon Thorvald",
        "Ilesa Ecaz",
        "Tessia Vernius"
    ]

    # Process each worksheet
    for sheet_name, display_name in resource_sheets.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Skipping {sheet_name} (not found)")
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        # Use "Card Name" for all sheets
        name_col = "Card Name"
        if name_col not in headers and headers:
            name_col = headers[0]

        merakon_col = "Count in Merakon's House Blend"
        merakon_idx = headers.index(merakon_col) if merakon_col in headers else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            resource_name = str(row_dict.get(name_col, '')).strip()
            if not resource_name:
                continue

            source = str(row_dict.get("Source", "Imperium")).strip()

            # Normalize Base to Imperium
            if source == "Base":
                source = "Imperium"

            resource_name_with_source = f"{resource_name} ({source})"

            # Get count from Excel
            merakon_count = 0
            if merakon_idx is not None:
                merakon_val = row_dict.get(merakon_col, 0)
                if merakon_val and str(merakon_val).strip() not in ['', '0', '0.0']:
                    try:
                        merakon_count = int(float(merakon_val))
                    except (ValueError, TypeError):
                        merakon_count = 0

            # Add based on Excel count
            if merakon_count > 0:
                for _ in range(merakon_count):
                    merakon_resources[display_name].append(resource_name_with_source)
            # Special handling for Leaders - add Uprising and specified leaders
            elif sheet_name == 'Leader':
                if resource_name in uprising_leaders or resource_name in other_leaders:
                    merakon_resources[display_name].append(resource_name_with_source)
            # Special handling for Reserve - add all Uprising ones
            elif sheet_name == 'Reserve' and source == "Uprising":
                # Get count from Count column
                count = row_dict.get("Count") or 1
                try:
                    item_count = int(float(count)) if count else 1
                except (ValueError, TypeError):
                    item_count = 1
                for _ in range(item_count):
                    merakon_resources[display_name].append(resource_name_with_source)
            # Special handling for Starter - add all Uprising ones
            elif sheet_name == 'Starter' and source == "Uprising":
                count = row_dict.get("Count") or row_dict.get("Count per Player") or 1
                try:
                    item_count = int(float(count)) if count else 1
                except (ValueError, TypeError):
                    item_count = 1
                for _ in range(item_count):
                    merakon_resources[display_name].append(resource_name_with_source)
            # Special handling for Conflict - add all Uprising ones
            elif sheet_name == 'Conflict' and source == "Uprising":
                count = row_dict.get("Count") or 1
                try:
                    item_count = int(float(count)) if count else 1
                except (ValueError, TypeError):
                    item_count = 1
                for _ in range(item_count):
                    merakon_resources[display_name].append(resource_name_with_source)
            # Special handling for Contracts - add all Uprising ones
            elif sheet_name == 'Contracts' and source == "Uprising":
                count = row_dict.get("Count") or 1
                try:
                    item_count = int(float(count)) if count else 1
                except (ValueError, TypeError):
                    item_count = 1
                for _ in range(item_count):
                    merakon_resources[display_name].append(resource_name_with_source)

    # Add manual leaders that might not be in Excel
    for leader_name, leader_source in manual_leaders.items():
        leader_with_source = f"{leader_name} ({leader_source})"
        if leader_with_source not in merakon_resources['Leaders']:
            merakon_resources['Leaders'].append(leader_with_source)

    # Create Merakon's blend (uses Uprising board)
    if any(merakon_resources.values()):
        filepath = blends_dir / "Merakons_House_Blend.md"
        create_multi_resource_blend_file(filepath, "Merakon's House Blend", merakon_resources,
                                        "", board="uprising")
        total = sum(len(items) for items in merakon_resources.values())
        print(f"✓ Created Merakon's House Blend with {total} total items")

        # Print summary
        for resource_type, items in merakon_resources.items():
            if items:
                print(f"  {resource_type}: {len(items)} items")


def create_multi_resource_blend_file(filepath, blend_name, resources_by_type, description="", board="imperium", additional_boards=None):
    """Create a blend file with multiple resource types in simplified format."""
    md = f"# {blend_name}\n\n"

    # Add board selection at the top
    md += f"## Board\n\n"
    md += f"- Main Board: {board}\n"
    if additional_boards:
        md += f"- Additional Boards: {', '.join(additional_boards)}\n"
    md += "\n"

    if description:
        md += f"*{description}*\n\n"

    total_items = sum(len(items) for items in resources_by_type.values())
    md += f"**Total Items:** {total_items}\n\n"

    # Add each resource type section
    for resource_type, items in resources_by_type.items():
        if not items:
            continue

        md += f"## {resource_type}\n\n"

        # Count occurrences
        item_counts = Counter(items)

        # Sort by name and output
        for item_name in sorted(item_counts.keys()):
            count = item_counts[item_name]
            if count == 1:
                md += f"- {item_name}\n"
            else:
                md += f"- {count}× {item_name}\n"

        md += "\n"

    md += "---\n*Generated by Dune Imperium Blend Builder*\n"

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(md)


if __name__ == "__main__":
    create_merakon_blend()
    print("\n✓ Merakon's House Blend created successfully!")

