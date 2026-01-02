#!/usr/bin/env python3
"""
Generate resources.json from Excel spreadsheet.
Run this whenever the Excel file is updated.
"""
import openpyxl
import json
from pathlib import Path


def generate_resources_json():
    """Load all resource types from Excel and save as JSON."""
    excel_path = Path(__file__).parent / "Dune_Imperium_Card_Inventory.xlsx"

    if not excel_path.exists():
        raise FileNotFoundError(f"Could not find: {excel_path}")

    print(f"Loading {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    all_resources = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        resources = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            row_dict = dict(zip(headers, row))

            # Get the name from first column
            name_col = headers[0]
            resource_name = str(row_dict.get(name_col, '')).strip()
            if not resource_name:
                continue

            # Normalize card names: replace (Base) with (Imperium)
            resource_name = resource_name.replace("(Base)", "(Imperium)")

            # Get source early to check for duplication
            source = str(row_dict.get('Source', 'Imperium')).strip()
            if source == 'Base':
                source = 'Imperium'

            # Check if the card name already has the source in parentheses
            source_suffix = f"({source})"
            if resource_name.endswith(source_suffix):
                resource_name = resource_name[:-len(source_suffix)].strip()

            # Create resource object
            resource = {
                'resource_type': sheet_name.lower(),
                'name': resource_name,
                'selected': 0
            }

            # Add all columns as properties
            for key, value in row_dict.items():
                if key and value is not None and key != name_col:
                    col_key = key.lower().replace(' ', '_').replace('-', '_')
                    resource[col_key] = str(value) if not isinstance(value, (int, float)) else value

            # Add source/set mapping for color coding
            source = resource.get('source', 'Imperium')
            if source == 'Base':
                source = 'Imperium'
                resource['source'] = 'Imperium'

            card_set_mapping = {
                "Imperium": "base",
                "Base": "base",
                "Rise of Ix": "ix",
                "Ix": "ix",
                "Immortality": "immortality",
                "Uprising": "uprising",
                "Bloodlines": "bloodlines",
                "Promo": "promo"
            }
            resource['card_set'] = card_set_mapping.get(source, str(source).lower() if source else 'base')

            resources.append(resource)

        all_resources[sheet_name.lower()] = resources
        print(f"Loaded {len(resources)} items from {sheet_name}")

    # Write JSON file
    output_path = Path(__file__).parent / 'resources.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_resources, f, indent=2)

    print(f"\n✅ Generated {output_path}")
    print(f"Total resource types: {len(all_resources)}")
    total_items = sum(len(resources) for resources in all_resources.values())
    print(f"Total items: {total_items}")


if __name__ == '__main__':
    generate_resources_json()

