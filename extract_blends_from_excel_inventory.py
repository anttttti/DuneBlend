#!/usr/bin/env python3
"""
Regenerate all blend files in the new simplified format
Format: "- count× name" for all resource types
"""
import json
import openpyxl
from pathlib import Path
from collections import Counter
from urllib.parse import quote

def generate_dune_card_hub_url(card_name, resource_type, expansion):
    """Generate a dunecardshub.com search URL for a card."""
    # Map resource types to dunecardshub.com types
    type_mapping = {
        'Imperium Cards': 'Imperium',
        'Intrigue Cards': 'Intrigue',
        'Reserve Cards': 'Reserve',
        'Tech Tiles': 'Tech',
        'Contracts': 'Contract',
        'Leaders': 'Leader',
        'Starter Cards': 'Starter',
        'Conflict Cards': 'Conflict',
        'Sardaukar': 'Sardaukar',
        'Tleilax Cards': 'Tleilax'
    }

    # Clean card name (remove count prefix and source suffix)
    clean_name = card_name
    if '×' in clean_name:
        clean_name = clean_name.split('×', 1)[1].strip()
    if '(' in clean_name and ')' in clean_name:
        clean_name = clean_name.rsplit('(', 1)[0].strip()

    card_type_param = type_mapping.get(resource_type, resource_type)
    search_term = quote(clean_name.lower())
    expansion_param = quote(expansion)

    url = f"https://dunecardshub.com/?search={search_term}&types={card_type_param}"
    if expansion and expansion != 'Unknown':
        url += f"&expansions={expansion_param}"

    return url

def get_starter_cards_for_source(source):
    """Get all starter cards for a specific source (Imperium or Uprising)."""
    excel_path = Path(__file__).parent / "Dune_Imperium_Card_Inventory.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if 'Starter' not in wb.sheetnames:
        return []

    ws = wb['Starter']
    headers = [cell.value for cell in ws[1]]
    name_col = headers[0] if headers else "Card Name"

    starter_cards = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        resource_name = str(row_dict.get(name_col, '')).strip()
        if not resource_name:
            continue

        # Get source and normalize
        card_source = str(row_dict.get("Source", "Imperium")).strip()
        if card_source == "Base":
            card_source = "Imperium"

        # Only include cards from the specified source
        if card_source != source:
            continue

        # Normalize name
        resource_name = resource_name.replace('(Base)', '(Imperium)')

        # Remove source suffix if present
        source_suffix = f"({card_source})"
        if resource_name.endswith(source_suffix):
            resource_name = resource_name[:-len(source_suffix)].strip()

        # Add with source
        resource_name_with_source = f"{resource_name} ({card_source})"

        # Get count
        count = row_dict.get("Count") or row_dict.get("Count per Player") or 1
        try:
            item_count = int(float(count)) if count else 1
        except (ValueError, TypeError):
            item_count = 1

        # Add the appropriate number of copies
        for _ in range(item_count):
            starter_cards.append(resource_name_with_source)

    return starter_cards

def regenerate_all_blends():
    excel_path = Path(__file__).parent / "Dune_Imperium_Card_Inventory.xlsx"
    blends_dir = Path(__file__).parent / "blends"
    blends_dir.mkdir(exist_ok=True)

    print(f"Loading {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Resource type configurations
    resource_sheets = {
        'Imperium': 'Imperium Cards',
        'Intrigue': 'Intrigue Cards',
        'Tleilax': 'Tleilax Cards',
        'Reserve': 'Reserve Cards',
        'Tech': 'Tech Tiles',
        'Contracts': 'Contracts',
        'Leader': 'Leaders',
        'Sardaukar': 'Sardaukar',
        'Starter': 'Starter Cards',
        'Conflict': 'Conflict Cards'
    }

    # For custom blends (Merakon and TragicJonson)
    merakon_resources = {sheet: [] for sheet in resource_sheets.values()}
    tragic_resources = {sheet: [] for sheet in resource_sheets.values()}

    # Merakon's special leader requirements
    merakon_manual_leaders = {
        '"Princess" Yuna Moritani': 'Rise of Ix',
        'Glossu "The Beast" Rabban': 'Imperium'
    }

    merakon_uprising_leaders = [
        "Feyd-Rautha Harkonnen",
        "Gurney Halleck",
        "Lady Amber Metulli",
        "Lady Jessica",
        "Lady Margot Fenring",
        "Muad'Dib",
        "Princess Irulan",
        "Shaddam Corrino IV",
        "Staban Tuek"
    ]

    merakon_other_leaders = [
        "Archduke Armand Ecaz",
        "Baron Vladimir Harkonnen",
        "Count Ilban Richese",
        "Countess Ariana Thorvald",
        "Earl Memnon Thorvald",
        "Ilesa Ecaz",
        "Tessia Vernius"
    ]

    # Process each worksheet
    for sheet_name, display_name in resource_sheets.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Skipping {sheet_name} (not found)")
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        # Find name column (first column usually)
        name_col = headers[0] if headers else "Card Name"

        # Find blend columns if they exist
        merakon_col = "Count in Merakon's House Blend"
        tragic_col = "Count in TragicJonson's House Blend"

        merakon_idx = headers.index(merakon_col) if merakon_col in headers else None
        tragic_idx = headers.index(tragic_col) if tragic_col in headers else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            # Properly handle None values - convert to empty string, not "None"
            resource_name_raw = row_dict.get(name_col, '')
            if resource_name_raw is None:
                continue
            resource_name = str(resource_name_raw).strip()
            if not resource_name or resource_name.upper() == 'NONE':
                continue

            # Skip Intrigue cards with Twisted = X
            if sheet_name == 'Intrigue':
                twisted = row_dict.get('Twisted', '')
                if twisted and str(twisted).strip().upper() == 'X':
                    continue

            # Normalize card names: replace (Base) with (Imperium)
            resource_name = resource_name.replace('(Base)', '(Imperium)')

            # Get source
            source = str(row_dict.get("Source", "Imperium")).strip()
            # Normalize Base to Imperium
            if source == "Base":
                source = "Imperium"

            # Check if the card name already has the source in parentheses
            # If so, remove it to avoid duplication
            source_suffix = f"({source})"
            if resource_name.endswith(source_suffix):
                # Remove the suffix from the name
                resource_name = resource_name[:-len(source_suffix)].strip()

            # Now add the source in parentheses
            resource_name_with_source = f"{resource_name} ({source})"

            # Add to Merakon's blend if count exists OR special handling
            merakon_count = 0
            if merakon_idx is not None:
                merakon_val = row_dict.get(merakon_col, 0)
                if merakon_val and str(merakon_val).strip() not in ['', '0', '0.0']:
                    try:
                        merakon_count = int(float(merakon_val))
                    except (ValueError, TypeError):
                        merakon_count = 0

            # Add based on Excel count
            if merakon_count > 0:
                for _ in range(merakon_count):
                    merakon_resources[display_name].append(resource_name_with_source)
            # Special handling for Leaders - add Uprising and specified leaders
            elif sheet_name == 'Leader':
                if resource_name in merakon_uprising_leaders or resource_name in merakon_other_leaders:
                    merakon_resources[display_name].append(resource_name_with_source)
            # Special handling for Reserve, Starter, Conflict, Contracts - add all Uprising ones
            elif sheet_name in ['Reserve', 'Starter', 'Conflict', 'Contracts'] and source == "Uprising":
                # Get count from appropriate column
                if sheet_name == 'Starter':
                    count = row_dict.get("Count") or row_dict.get("Count per Player") or 1
                else:
                    count = row_dict.get("Count") or 1
                try:
                    item_count = int(float(count)) if count else 1
                except (ValueError, TypeError):
                    item_count = 1
                for _ in range(item_count):
                    merakon_resources[display_name].append(resource_name_with_source)

            # Add to TragicJonson's blend if count exists
            if tragic_idx is not None:
                tragic_count = row_dict.get(tragic_col, 0)
                if tragic_count and str(tragic_count).strip() not in ['', '0', '0.0']:
                    try:
                        count = int(float(tragic_count))
                        if count > 0:
                            for _ in range(count):
                                tragic_resources[display_name].append(resource_name_with_source)
                    except (ValueError, TypeError):
                        pass

    # Add manual leaders to Merakon's blend
    for leader_name, leader_source in merakon_manual_leaders.items():
        leader_with_source = f"{leader_name} ({leader_source})"
        if leader_with_source not in merakon_resources['Leaders']:
            merakon_resources['Leaders'].append(leader_with_source)

    # Create Merakon's blend (uses Uprising board)
    if any(merakon_resources.values()):
        filepath = blends_dir / "Merakons_House_Blend.md"
        create_multi_resource_blend_file(
            filepath, "Merakon's House Blend", merakon_resources,
            description="https://boardgamegeek.com/thread/3213458/merakons-house-blend",
            board="uprising",
            leader_selection="Deal four leaders to each player. Everyone picks a leader simultaneously."
        )
        total = sum(len(items) for items in merakon_resources.values())
        print(f"✓ Created Merakon's House Blend with {total} total items")

    # Create TragicJonson's blend (uses Uprising board)
    if any(tragic_resources.values()):
        filepath = blends_dir / "TragicJonsons_House_Blend.md"
        create_multi_resource_blend_file(
            filepath, "TragicJonson's House Blend", tragic_resources,
            description="https://observablehq.com/@mrcorvus/dune-imperium-deck-builder",
            board="uprising"
        )
        total = sum(len(items) for items in tragic_resources.values())
        print(f"✓ Created TragicJonson's House Blend with {total} total items")

    # Create Base Imperium and Base Uprising
    create_base_blends(wb, resource_sheets)


def create_base_blends(wb, resource_sheets):
    """Create Base Imperium and Base Uprising blends."""
    # Load resources.json to get resource IDs for synonym handling
    resources_json_path = Path(__file__).parent / 'resources.json'
    with open(resources_json_path, 'r', encoding='utf-8') as f:
        all_resources = json.load(f)

    base_imperium_resources = {}
    base_uprising_resources = {}

    for sheet_name, display_name in resource_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue

        # Skip Sardaukar - they are physical tokens that come with the game regardless
        # Skip Tech - not part of base game card selections
        # Skip Contracts - only in Imperium base game, not in base blends
        if sheet_name in ['Sardaukar', 'Tech', 'Contracts']:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        name_col = headers[0] if headers else "Card Name"

        base_imperium_items = []
        base_uprising_items = []

        # Track which synonyms we've already added (to avoid adding them multiple times)
        added_synonyms = set()

        # Build a lookup of resources by name and source from resources.json
        resource_type_key = sheet_name.lower()
        resource_lookup = {}
        if resource_type_key in all_resources:
            for resource in all_resources[resource_type_key]:
                name = resource.get('name', '')
                source = resource.get('source', 'Imperium')
                resource_id = resource.get('resource_id', 0)
                key = f"{name}|{source}"
                if key not in resource_lookup:
                    resource_lookup[key] = []
                resource_lookup[key].append(resource_id)

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            # Properly handle None values - convert to empty string, not "None"
            resource_name_raw = row_dict.get(name_col, '')
            if resource_name_raw is None:
                continue
            resource_name = str(resource_name_raw).strip()
            if not resource_name or resource_name.upper() == 'NONE':
                continue

            # Normalize card names: replace (Base) with (Imperium)
            resource_name = resource_name.replace('(Base)', '(Imperium)')

            source = str(row_dict.get("Source", "Imperium")).strip()
            # Normalize Base to Imperium
            if source == "Base":
                source = "Imperium"

            # Check if the card name already has the source in parentheses
            # If so, remove it to avoid duplication
            source_suffix = f"({source})"
            if resource_name.endswith(source_suffix):
                # Remove the suffix from the name
                resource_name = resource_name[:-len(source_suffix)].strip()

            # Handle different count column names (Starter uses "Count per Player")
            count = row_dict.get("Count") or row_dict.get("Count per Player") or 1

            try:
                item_count = int(float(count)) if count else 1
            except (ValueError, TypeError):
                item_count = 1

            # Check if this is a synonym (multiple resources with same name+source)
            lookup_key = f"{resource_name}|{source}"
            resource_ids = resource_lookup.get(lookup_key, [])

            if len(resource_ids) > 1:
                # This is a synonym - add each variant once (only if not already added)
                if lookup_key not in added_synonyms:
                    added_synonyms.add(lookup_key)
                    for idx in range(1, len(resource_ids) + 1):
                        resource_with_id = f"{resource_name} #{idx} ({source})"
                        if source == "Imperium":
                            base_imperium_items.append(resource_with_id)
                        elif source == "Uprising":
                            base_uprising_items.append(resource_with_id)
            else:
                # Not a synonym - use regular name with source
                resource_name_with_source = f"{resource_name} ({source})"
                if source == "Imperium":
                    for _ in range(item_count):
                        base_imperium_items.append(resource_name_with_source)
                elif source == "Uprising":
                    for _ in range(item_count):
                        base_uprising_items.append(resource_name_with_source)

        if base_imperium_items:
            base_imperium_resources[display_name] = base_imperium_items
        if base_uprising_items:
            base_uprising_resources[display_name] = base_uprising_items

    # Save Base Imperium (uses imperium board)
    if base_imperium_resources:
        filepath = Path(__file__).parent / "blends" / "Base_Imperium.md"
        create_multi_resource_blend_file(filepath, "Base Imperium", base_imperium_resources,
                                        "Dune: Imperium base game", board="imperium")
        total = sum(len(items) for items in base_imperium_resources.values())
        print(f"✓ Created Base Imperium with {total} total items")

    # Save Base Uprising (uses uprising board)
    if base_uprising_resources:
        filepath = Path(__file__).parent / "blends" / "Base_Uprising.md"
        create_multi_resource_blend_file(filepath, "Base Uprising", base_uprising_resources,
                                        "Dune: Imperium - Uprising base game", board="uprising")
        total = sum(len(items) for items in base_uprising_resources.values())
        print(f"✓ Created Base Uprising with {total} total items")


def create_multi_resource_blend_file(filepath, blend_name, resources_by_type, description="", board="imperium", additional_boards=None, leader_selection="", house_rules=""):
    """Create a blend file with multiple resource types in simplified format."""
    # Auto-add Starter cards based on board
    if 'Starter Cards' not in resources_by_type or not resources_by_type['Starter Cards']:
        # Add starter cards based on board
        starter_source = "Imperium" if board == "imperium" else "Uprising"
        resources_by_type['Starter Cards'] = get_starter_cards_for_source(starter_source)

    md = f"# {blend_name}\n\n"

    # Add Overview section if any overview fields are provided
    if description or leader_selection or house_rules:
        md += "## Overview\n\n"
        if description:
            md += "### Description\n\n"
            md += f"{description}\n\n"
        if leader_selection:
            md += "### Leader Selection\n\n"
            md += f"{leader_selection}\n\n"
        if house_rules:
            md += "### House Rules\n\n"
            md += f"{house_rules}\n\n"

    # Add board selection
    md += f"## Board\n\n"
    md += f"- Main Board: {board}\n"
    if additional_boards:
        md += f"- Additional Boards: {', '.join(additional_boards)}\n"
    md += "\n"


    total_items = sum(len(items) for items in resources_by_type.values())
    md += f"**Total Items:** {total_items}\n\n"

    # Add each resource type section
    for resource_type, items in resources_by_type.items():
        if not items:
            continue

        md += f"## {resource_type}\n\n"

        # Separate synonym items (with #) from regular items
        synonym_items = []
        regular_items = []
        for item in items:
            if ' #' in item and item.count('(') > 0:
                # This is a synonym item - keep it separate
                synonym_items.append(item)
            else:
                regular_items.append(item)

        # Count occurrences for regular items only
        item_counts = Counter(regular_items)

        # Sort by name and output regular items
        for item_name in sorted(item_counts.keys()):
            count = item_counts[item_name]
            if count == 1:
                md += f"- {item_name}\n"
            else:
                # Multiple copies - use count× format
                md += f"- {count}× {item_name}\n"

        # Output synonym items individually (no grouping)
        for item_name in sorted(set(synonym_items)):
            # Count how many of this specific synonym variant
            count = synonym_items.count(item_name)
            if count == 1:
                md += f"- {item_name}\n"
            else:
                md += f"- {count}× {item_name}\n"

        md += "\n"

    md += "---\n*Generated by Dune Imperium Blend Builder*\n"

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(md)


if __name__ == '__main__':
    regenerate_all_blends()
    print("\n✓ All blend files regenerated with all resource types!")

